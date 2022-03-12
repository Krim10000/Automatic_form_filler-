
#import os
import pyautogui 
import time
import cv2 # vision
import openpyxl # excel
import clipboard

from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("File.xlsx") 
ws = wb.active

#cell_obj = ws.cell(row = 1, column = 1) 
  
# Print value of cell object  
# using the value attribute 
#print(cell_obj.value)




#13466623-4


print ("###############")
print ("")

print("Bienvenida")
print ("Ingrese rut del usuario:")
usuario = input()
###  BUSCAR EN EXCEL

#Rut Empleador
#Vloookup
for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:#TAB#
        for j in range(i, ws.max_column):
            RUTE=(ws.cell(row=i, column=23).value)
RUTE=str(RUTE)            
print("Rut empresa "+ RUTE)            


#Razon social

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            RSOC= (ws.cell(row=i, column=24).value)
RSOC=str(RSOC)
RSOC= RSOC.upper()
print("Razon social "+RSOC)          


#CALLE EMPRESA

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            CAEMP= (ws.cell(row=i, column=27).value)
CAEMP=str(CAEMP.upper())            
print("Calle empresa "+CAEMP)    

#NUM EMPRESA

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            NEMP= (ws.cell(row=i, column=28).value)
NEMP=str(NEMP)            
print("Numero empresa "+NEMP)

#COMUNA EMPRESA

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            COEMP= (ws.cell(row=i, column=25).value)
COEMP=str(COEMP)            
print("Comuna empresa "+COEMP)


#ACERCAMIENTO EMPRESA

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            AEMP= (ws.cell(row=i, column=29).value)
AEMP=str(AEMP)
if AEMP == "None":
    AEMP = ""
print("Acercamiento empresa "+AEMP)


#TELEFONO EMPRESA

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            TEMP= (ws.cell(row=i, column=31).value)
TEMP=str(TEMP)            
print("Telefono empresa "+TEMP)



#CALLE TRABAJADOR

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            CATAB= (ws.cell(row=i, column=12).value)
CATAB=str(CATAB)            
print("Calle trabajador "+CATAB)


#COMUNA TRABAJADOR

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            COTAB= (ws.cell(row=i, column=10).value)
COTAB=str(COTAB)            
print("COMUNA trabajador "+COTAB)


#NUMERO TRABAJADOR

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            NTAB= (ws.cell(row=i, column=13).value)
NTAB=str(NTAB)            
print("NUMERO trabajador "+NTAB)

#TELEFONO TRABAJADOR

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            TTAB= (ws.cell(row=i, column=18).value)
TTAB=str(TTAB)            
print("TELEFONO trabajador "+TTAB)

#CORREO TRABAJADOR

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            CTAB= (ws.cell(row=i, column=17).value)
CTAB=str(CTAB)            
print("CORREO trabajador "+CTAB)


#OTRA FUNCION TRABAJADOR

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            OTFUN= (ws.cell(row=i, column=21).value)
OTFUN=str(OTFUN)            
print("OTRA FUNCION trabajador "+OTFUN)

#INICIO TRABAJADOR

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            IREL= (ws.cell(row=i, column=19).value)
IREL=str(IREL)            
print("INICIO trabajador "+IREL)

#FIN TRABAJADOR

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            FREL= (ws.cell(row=i, column=20).value)
FREL=str(FREL)            
print("FIN trabajador "+FREL)

#REGIMEN PREVISIONAL TRABAJADOR

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            REGPREV= (ws.cell(row=i, column=16).value)
REGPREV=str(REGPREV[0])
if REGPREV != ("A" or "I"):
    REGPREV = "N"
print("PREVISIONAL trabajador "+REGPREV)



#REGIMEN SALUD TRABAJADOR

for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            REGSAL= (ws.cell(row=i, column=15).value)
REGSAL=str(REGSAL[0])
if REGSAL != ("F" or "I"):
    REGSAL = "N"
print("SALUD trabajador "+REGSAL)


#JORNADA


for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            JOR= (ws.cell(row=i, column=22).value)
JOR=str(JOR.upper())
print("JORNADA trabajador "+JOR)



if "BISE" in JOR:
    JOR = int(1)
    

elif "COMP" in JOR:
    JOR = int(2)
    

elif "CONV" in JOR:
    JOR = int(3)
    

elif "EXCE" in JOR:
    JOR = int(4)
    

elif "PAR" in JOR:
    JOR = int(5)
    

else:
    JOR = int(0)

###Correo empleador     




for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            CORREO= (ws.cell(row=i, column=30).value)
CORREO=str(CORREO)            
print("CORREO empleador "+CORREO)


###Observaciones    




for i in range(1,ws.max_row):
    if ws.cell(row=i, column=4).value == usuario:
        for j in range(i, ws.max_column):
            COMEN= (ws.cell(row=i, column=52).value)
COMEN=str(COMEN)            
print("COMENTARIO empleado "+COMEN)

############################################################################################
### INICIO PAGINA WEB
while True:
    
    R= pyautogui.locateCenterOnScreen("INI.png", confidence=0.8)

    print(R)

    if R != None:
        break

    

###########################################################

### ESCRIBIR EN PAGINA

#INICIO RECLAMO

R= pyautogui.locateCenterOnScreen("INI.png")

print(R)
x, y = pyautogui.locateCenterOnScreen("INI.png")


pyautogui.click(x, y)

if usuario == 10:
    for x in range (9):
        pyautogui.typewrite(usuario[x])
        
else:
    for x in range (8):
        pyautogui.typewrite(usuario[x])


pyautogui.typewrite(usuario[-1])
time.sleep(1)
pyautogui.press("enter")
time.sleep(1)
pyautogui.click(x, y+80)


time.sleep(3)
#13466623-4




####### BAJAR!

while True:
    
    R= pyautogui.locateCenterOnScreen("inicio.png", confidence=0.8)

    print(R)
    pyautogui.press('down')

    if R != None:
        break

#RAE
try:
    R= pyautogui.locateCenterOnScreen("RAE.png")

    print(R)
    x, y = pyautogui.locateCenterOnScreen("RAE.png")


    pyautogui.click(x, y)

    pyautogui.typewrite("c", interval=0)

    pyautogui.typewrite("c", interval=0)

    

except:
    print("RAE error")





#Utilizar este domicilio como domicilio de Faena o Prestación de Servicios 
try:
    R= pyautogui.locateCenterOnScreen("utilizar.png")

    print(R)
    x, y = pyautogui.locateCenterOnScreen("utilizar.png")


    pyautogui.click(x-15, y)

except:
    print("utilizar error")
 
###RUT EMPRESA
try: 
    R= pyautogui.locateCenterOnScreen("RUT_EMP.png")

    print(R)
    x, y = pyautogui.locateCenterOnScreen("RUT_EMP.png")

    pyautogui.click(x, y)

    ##12875142-4
    ##12.875.142-4
    ##96.599.510-2 2+3+3+2 =10
    #92.625.292.925.212-0
    ##96599510-2
    ##96599510-2

    
    if len(RUTE)==10:
        for x in range (8):
            pyautogui.typewrite(RUTE[x])

    else:
        for x in range (7):
            pyautogui.typewrite(RUTE[x])

    pyautogui.typewrite(RUTE[-1])

   

    
except:
    print("RUTE error")

###RAZON SOCIAL
try:
    R= pyautogui.locateCenterOnScreen("RAZON_SOCIAL.png")

    print(R)
    x, y = pyautogui.locateCenterOnScreen("RAZON_SOCIAL.png")


    pyautogui.click(x+30, y)

    pyautogui.typewrite(RSOC, interval=0)

except:
    print("RSOC error")

#CALLE EMPRESA

try:
    R= pyautogui.locateCenterOnScreen("CALLE_EMP.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("CALLE_EMP.png",confidence =0.8)


    pyautogui.click(x, y)

    pyautogui.typewrite(CAEMP, interval=0)

except:
    print("CAEMP error")


#NUMERO EMPRESA
try:
    R= pyautogui.locateCenterOnScreen("NUM_EMP.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("NUM_EMP.png",confidence =0.8)


    pyautogui.click(x, y)

    pyautogui.typewrite(NEMP, interval=0)

except:
    print("NEMP error")

#comuna EMPRESA
try:
    R= pyautogui.locateCenterOnScreen("COM_EMP.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("COM_EMP.png",confidence =0.8)


    pyautogui.click(x, y)

    pyautogui.typewrite(COEMP, interval=0)
    
    time.sleep(1)
    pyautogui.click(x, y+40)
    

except:
    print("COEMP error")
    
#ACERCAMIENTO EMPRESA
try:
    R= pyautogui.locateCenterOnScreen("ACE_EMP.png")

    print(R)
    x, y = pyautogui.locateCenterOnScreen("ACE_EMP.png")


    pyautogui.click(x, y)

    pyautogui.typewrite(AEMP, interval=0)
except:
    print("AEMP error")

#TELEFONO EMPRESA

try:
    R= pyautogui.locateCenterOnScreen("T_EMP.png")

    print(R)
    x, y = pyautogui.locateCenterOnScreen("T_EMP.png")


    pyautogui.click(x, y)

    pyautogui.typewrite(TEMP, interval=0)
except:
    print("TEMP error")

#####subir
pyautogui.click(100, 500)    
pyautogui.press('home')

##### RECLAMANTE

try:
    R= pyautogui.locateCenterOnScreen("RECLAMANTE.png")

    print(R)
    x, y = pyautogui.locateCenterOnScreen("RECLAMANTE.png", confidence =0.8)


    pyautogui.click(x, y)

    pyautogui.typewrite(TEMP, interval=0)
except:
    print("RECLAMANTE error")
time.sleep(2)
####### BAJAR!

while True:
    
    R= pyautogui.locateCenterOnScreen("INFOCON.png", confidence=0.8)

    print(R)
    pyautogui.press('down')

    if R != None:
        break

##### DOMICILIO RECLAMANTE


#CALLE 

try:
    R= pyautogui.locateCenterOnScreen("CATAB.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("CATAB.png",confidence =0.8)


    pyautogui.click(x, y)

    pyautogui.typewrite(CATAB, interval=0)

except:
    print("CATAB error")


#NUMERO 
try:
    R= pyautogui.locateCenterOnScreen("NUM_EMP.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("NUM_EMP.png",confidence =0.8)


    pyautogui.click(x, y)

    pyautogui.typewrite(NTAB, interval=0)

except:
    print("NTAB error")

#COMUNA 
try:
    R= pyautogui.locateCenterOnScreen("COM_EMP.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("COM_EMP.png",confidence =0.8)


    pyautogui.click(x, y)

    pyautogui.typewrite(COTAB, interval=0)
    
    time.sleep(1)
    pyautogui.click(x, y+40)
    

except:
    print("COTAB error")
    


##### CONTACTO

### TELEFONO    
try:
    R= pyautogui.locateCenterOnScreen("T_EMP.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("T_EMP.png",confidence =0.8)

    pyautogui.click(x, y)

    pyautogui.typewrite(TTAB, interval=0)
    
except:
    print("TTAB error")
    
### CORREO

try:
    R= pyautogui.locateCenterOnScreen("CTAB.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("CTAB.png",confidence =0.8)

    pyautogui.click(x, y)

    pyautogui.typewrite(CTAB, interval=0)
    
except:
    print("CTAB error")
#############################    
### FIN PAGINA
pyautogui.click(100, 500)    
pyautogui.press('end')

##### INFO LAB

###FUNCION
try:
    R= pyautogui.locateCenterOnScreen("FUTAB.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("FUTAB.png",confidence =0.8)

    pyautogui.click(x, y)

    pyautogui.typewrite("o", interval=0)
    pyautogui.typewrite("o", interval=0)

    pyautogui.press("enter")
    
except:
    print("FUTAB error")

###OTRA FUNCION

try:
    R= pyautogui.locateCenterOnScreen("OTFUN.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("OTFUN.png",confidence =0.8)

    pyautogui.click(x, y)

    pyautogui.typewrite(OTFUN, interval=0)
    
except:
    print("OTFUN error")
    

###INICIO 

try:
    R= pyautogui.locateCenterOnScreen("IREL.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("IREL.png",confidence =0.8)

    pyautogui.click(x, y)

    pyautogui.typewrite(IREL[0], interval=0)
    pyautogui.typewrite(IREL[1], interval=0)
    pyautogui.typewrite(IREL[3], interval=0)
    pyautogui.typewrite(IREL[4], interval=0)
    pyautogui.typewrite(IREL[6], interval=0)
    pyautogui.typewrite(IREL[7], interval=0)
    pyautogui.typewrite(IREL[8], interval=0)
    pyautogui.typewrite(IREL[9], interval=0)
except:
    print("IREL error")


###FIN 

try:
    R= pyautogui.locateCenterOnScreen("FREL.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("FREL.png",confidence =0.8)

    pyautogui.click(x, y)

    pyautogui.typewrite(FREL[0], interval=0)
    pyautogui.typewrite(FREL[1], interval=0)
    pyautogui.typewrite(FREL[3], interval=0)
    pyautogui.typewrite(FREL[4], interval=0)
    pyautogui.typewrite(FREL[6], interval=0)
    pyautogui.typewrite(FREL[7], interval=0)
    pyautogui.typewrite(FREL[8], interval=0)
    pyautogui.typewrite(FREL[9], interval=0)
    
except:
    print("FREL error")

    
###RELACION  PREVISIONAL

try:
    R= pyautogui.locateCenterOnScreen("REGPREV.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("REGPREV.png",confidence =0.8)

    pyautogui.click(x, y)

    pyautogui.typewrite(REGPREV, interval=0)
    pyautogui.press("enter")
    
except:
    print("REGPREV error")


###RELACION  SALUD

try:
    R= pyautogui.locateCenterOnScreen("REGSAL.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("REGSAL.png",confidence =0.8)

    pyautogui.click(x, y)

    pyautogui.typewrite(REGSAL, interval=0)
    pyautogui.press("enter")
    
except:
    print("REGSAL error")


###JORNADA

try:
    R= pyautogui.locateCenterOnScreen("JORNADA.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("JORNADA.png",confidence =0.8)

    pyautogui.click(x, y)

      
 #########prodria haber usado un loop    

    if JOR == 1:
        pyautogui.press("down")
        
    elif JOR == 2:
        pyautogui.press("down")
        pyautogui.press("down")
        
    elif JOR == 3:
        pyautogui.press("down")
        pyautogui.press("down")
        pyautogui.press("down")
        
    elif JOR == 4:
        pyautogui.press("down")
        pyautogui.press("down")
        pyautogui.press("down")
        pyautogui.press("down")
        
    elif JOR == 5:
        pyautogui.press("down")
        pyautogui.press("down")
        pyautogui.press("down")
        pyautogui.press("down")
        pyautogui.press("down")
        
    
    pyautogui.press("enter")
    
    
    
except:
    print("JORNADA error")




### INICIO PAGINA
pyautogui.click(100, 200)    
pyautogui.press('home')

###Documentacion/conceptos

try:
    R= pyautogui.locateCenterOnScreen("docu.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("docu.png",confidence =0.8)

    pyautogui.click(x,y)
    
except:
    print("documentacion error")

time.sleep(3)


###INICIO 

try:
    R= pyautogui.locateCenterOnScreen("finicio.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("finicio.png",confidence =0.8)

    pyautogui.click(x, y)

    pyautogui.typewrite(IREL[0], interval=0)
    pyautogui.typewrite(IREL[1], interval=0)
    pyautogui.typewrite(IREL[3], interval=0)
    pyautogui.typewrite(IREL[4], interval=0)
    pyautogui.typewrite(IREL[6], interval=0)
    pyautogui.typewrite(IREL[7], interval=0)
    pyautogui.typewrite(IREL[8], interval=0)
    pyautogui.typewrite(IREL[9], interval=0)
except:
    print("IREL error")


###FIN 

try:
    R= pyautogui.locateCenterOnScreen("ftermino.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("ftermino.png",confidence =0.8)

    pyautogui.click(x, y)

    pyautogui.typewrite(FREL[0], interval=0)
    pyautogui.typewrite(FREL[1], interval=0)
    pyautogui.typewrite(FREL[3], interval=0)
    pyautogui.typewrite(FREL[4], interval=0)
    pyautogui.typewrite(FREL[6], interval=0)
    pyautogui.typewrite(FREL[7], interval=0)
    pyautogui.typewrite(FREL[8], interval=0)
    pyautogui.typewrite(FREL[9], interval=0)
    
except:
    print("FREL error")

### agregar conceptos
try:
    C= pyautogui.locateCenterOnScreen("concepto.png",confidence =0.8)

    print(C)
    x, y = pyautogui.locateCenterOnScreen("concepto.png",confidence =0.8)

    pyautogui.click(x,y)
    
except:
    print("concepto error")


A = pyautogui.locateCenterOnScreen("add.png",confidence =0.8)

#BIEN
pyautogui.press("a", interval=0)
pyautogui.press("enter", interval =1)
pyautogui.click(A)
print("Aporte seguro cesantia")

#mal
pyautogui.click(C)
pyautogui.press("f", interval=0)
pyautogui.press("enter", interval =1)
pyautogui.click(A)
print("Feriado legal/proporcional")

#BIEN
pyautogui.click(C)
pyautogui.press("i", interval=0)
pyautogui.press("enter", interval =1)
print("Indemnizacion falta de aviso previo")

pyautogui.click(A)

#BIEN
pyautogui.click(C)
pyautogui.press("f", interval=1)
pyautogui.press("f", interval=1)
pyautogui.press("f", interval=1)
pyautogui.press("enter", interval =1)
pyautogui.click(A)
print("Finiquito")

#BIEN
pyautogui.click(C)
pyautogui.press("i", interval=1)
pyautogui.press("i", interval=1)
pyautogui.press("i", interval=1)
pyautogui.press("enter", interval =1)
time.sleep(1)
print("Indemnizacion por años de servicio")
pyautogui.click(A)


#mal
pyautogui.click(C)
pyautogui.press("end", interval=1)
pyautogui.press("up", interval=1)
pyautogui.press("enter", interval =1)

time.sleep(1)
print("Remuneracion fija")
pyautogui.click(A)


x, y = pyautogui.locateCenterOnScreen("add.png",confidence =0.8)

pyautogui.click(x,y+50)

#una F +
#tres F +
#I +
#tres I+
#R +
#A+ listo



pyautogui.press("end", interval=0)



###Observaciones

try:
    R= pyautogui.locateCenterOnScreen("OBS.png",confidence =0.8)

    print(R)
    x, y = pyautogui.locateCenterOnScreen("OBS.png",confidence =0.8)
    pyautogui.click(x,y)


    comentario = CORREO +" || "+ COMEN
    comentario = comentario[0:255]

    clipboard.copy(comentario)
    pyautogui.hotkey('ctrl', 'v', interval = 0.15)
    #pyautogui.typewrite(comentario, interval=0)
    
except:
    print("Obs error")

    

print("")
print("")
print("")
print("TODO LISTO!")
print(usuario)

